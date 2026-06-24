"""
excel_writer.py
Output Excel: sintesi 474 + regolamento e UNO sheet di dettaglio per ogni check.
I dettagli usano la stessa base e la stessa logica (esenzioni incluse) dei check,
quindi il semaforo del dettaglio non puo' contraddire l'esito del check.
"""

import io
import datetime
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.analisi import CheckResult, BASE_LABEL

C_BLUE="00338D"; C_WHITE="FFFFFF"; C_BLACK="000000"; C_STRIPE="F2F2F2"
C_OK="C6EFCE"; C_WARN="FFEB9C"; C_ERR="FFC7CE"; C_BORDER="BFBFBF"; C_GRAY="D9D9D9"
C_ESENTE="BDD7EE"  # azzurro: riga esente dal limite
FONT_BODY="Arial"; FONT_LOGO="KPMG Logo"; FONT_BOLD="KPMG Bold"; SZ=8; SZ_HEAD=14

def _side(): return Side(style="thin", color=C_BORDER)
def _border(): s=_side(); return Border(left=s,right=s,top=s,bottom=s)

def _header_cell(ws,row,col,value,bg=C_BLUE):
    c=ws.cell(row,col,str(value))
    c.font=Font(name=FONT_BODY,size=SZ,bold=True,color=C_WHITE)
    c.fill=PatternFill("solid",start_color=bg)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=_border(); ws.row_dimensions[row].height=20

def _data_cell(ws,row,col,value,stripe=False,fmt=None,bg=None):
    c=ws.cell(row,col,value)
    c.font=Font(name=FONT_BODY,size=SZ,color=C_BLACK)
    c.alignment=Alignment(vertical="center"); c.border=_border()
    c.fill=PatternFill("solid",start_color=(bg or (C_STRIPE if stripe else C_WHITE)))
    if fmt: c.number_format=fmt
    ws.row_dimensions[row].height=12

def _kpmg_logo(ws,title):
    ws.sheet_view.showGridLines=False
    ws.column_dimensions["A"].width=2
    ws["B2"].value="kpmg"; ws["B2"].font=Font(name=FONT_LOGO,size=SZ_HEAD,color=C_BLUE)
    ws["B2"].alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[2].height=20
    ws["B3"].value=title; ws["B3"].font=Font(name=FONT_BOLD,size=SZ_HEAD,bold=True,color=C_BLUE)
    ws["B3"].alignment=Alignment(horizontal="left",vertical="center"); ws.row_dimensions[3].height=20
    ws.row_dimensions[4].height=6

def _auto_width(ws,start_col=2,min_w=8,max_w=60):
    for col in ws.iter_cols(min_col=start_col):
        w=min_w
        for cell in col:
            if cell.value: w=max(w,min(len(str(cell.value))+3,max_w))
        ws.column_dimensions[get_column_letter(col[0].column)].width=w

ESITO_COLOR={"OK":C_OK,"SFORAMENTO MAX":C_ERR,"SFORAMENTO EMITTENTE":C_ERR,
    "SOTTO MINIMO":C_WARN,"NON RILEVABILE":C_GRAY,"AVVISO":C_WARN}

HEADERS_474=["Norma","Check","Descrizione","Art./Par.","Limite MAX %","Limite MIN %",
    "Valore effettivo %","Esito","Scostamento pp","Dettaglio"]

def _write_check_sheet(ws,results,title,header_bg=C_BLUE):
    _kpmg_logo(ws,title)
    if not results:
        ws["B5"].value="Nessun check disponibile."
        ws["B5"].font=Font(name=FONT_BODY,size=SZ,color=C_BLACK); return
    for j,h in enumerate(HEADERS_474,2): _header_cell(ws,5,j,h,bg=header_bg)
    for i,r in enumerate(results,6):
        bg_row=ESITO_COLOR.get(r.esito,C_WHITE)
        vals=[r.norma,r.check_id,r.descrizione,r.articolo,r.limite_max_pct,r.limite_min_pct,
              r.valore_effettivo_pct,r.esito,r.scostamento_pp,r.dettaglio]
        ws.row_dimensions[i].height=12
        for j,val in enumerate(vals,2):
            c=ws.cell(i,j,val)
            c.font=Font(name=FONT_BODY,size=SZ,color=C_BLACK)
            c.alignment=Alignment(vertical="center"); c.border=_border()
            c.fill=PatternFill("solid",start_color=bg_row)
            if j in (6,7,8,10) and isinstance(val,(int,float)): c.number_format='0.00"%"'
    _auto_width(ws)

# ---------------------------------------------------------------------------
# DETTAGLIO PER SINGOLO CHECK
# ---------------------------------------------------------------------------
def _fmt_for(colname):
    n=str(colname).lower()
    if "%" in n: return '0.00"%"'
    if "eur" in n or "valore" in n: return '#,##0'
    return None

def _row_bg(rowvals, headers, meta, stripe):
    """Colore della riga del dettaglio, coerente con la logica del check."""
    if not meta: return C_STRIPE if stripe else C_WHITE
    pct_col=meta.get("pct_col")
    if not pct_col or pct_col not in headers: return C_STRIPE if stripe else C_WHITE
    es_col=meta.get("esente_col")
    if es_col and es_col in headers and bool(rowvals[headers.index(es_col)]):
        return C_ESENTE
    v=rowvals[headers.index(pct_col)]
    if v is None: return C_STRIPE if stripe else C_WHITE
    lim_max=meta.get("limite_max"); lim_min=meta.get("limite_min")
    warn=meta.get("warn_ratio",0.8)
    if lim_max is not None and v>lim_max: return C_ERR
    if lim_min is not None and v<lim_min: return C_WARN
    if lim_max is not None and v>warn*lim_max: return C_WARN
    return C_OK

def _write_detail_sheet(ws,res):
    _kpmg_logo(ws,f"Dettaglio check - {res.descrizione}")
    # --- banda riepilogo (aggancia il dettaglio all'esito del check) ---
    band=[("Check",res.check_id),("Articolo/Norma",res.articolo or res.norma),
          ("Limite MAX %",res.limite_max_pct),("Limite MIN %",res.limite_min_pct),
          ("Base di calcolo",f"{BASE_LABEL.get(res.base_calcolo,res.base_calcolo)} = €{res.base_valore:,.0f}"),
          ("Valore effettivo %",res.valore_effettivo_pct),("Esito",res.esito)]
    r=5
    for lbl,val in band:
        lc=ws.cell(r,2,lbl); lc.font=Font(name=FONT_BODY,size=SZ,bold=True); lc.border=_border()
        vc=ws.cell(r,3,val); vc.font=Font(name=FONT_BODY,size=SZ); vc.border=_border()
        if lbl=="Esito": vc.fill=PatternFill("solid",start_color=ESITO_COLOR.get(res.esito,C_WHITE))
        if lbl.endswith("%") and isinstance(val,(int,float)): vc.number_format='0.00"%"'
        ws.row_dimensions[r].height=13; r+=1

    df=res.dettaglio_df
    r+=1
    if df is None or len(df)==0:
        ws.cell(r,2,"Nessuna posizione rientrante in questo check.").font=Font(name=FONT_BODY,size=SZ,italic=True)
        _auto_width(ws); return

    headers=list(df.columns)
    meta=res.dettaglio_meta or {}
    head_row=r
    for j,h in enumerate(headers,2): _header_cell(ws,head_row,j,h)
    for i,(_,row) in enumerate(df.iterrows(),head_row+1):
        rowvals=list(row.values)
        bg=_row_bg(rowvals,headers,meta,stripe=(i%2==0))
        for j,(h,val) in enumerate(zip(headers,rowvals),2):
            if isinstance(val,bool): val="Sì" if val else "-"
            _data_cell(ws,i,j,val,fmt=_fmt_for(h),bg=bg)
    # legenda mini se ci sono esenti
    if meta.get("esente_col"):
        note_r=head_row+len(df)+2
        ws.cell(note_r,2,"Righe in azzurro = emittenti esenti dal limite (Titoli di Stato AAA): "
                "il loro valore non concorre al calcolo del check.").font=Font(name=FONT_BODY,size=SZ,italic=True)
    _auto_width(ws)

def _safe_sheet_name(base,used):
    name=re.sub(r"[\\/*?:\[\]]","_",f"D_{base}")[:31]
    k=name; n=1
    while k.lower() in used:
        suf=f"~{n}"; k=name[:31-len(suf)]+suf; n+=1
    used.add(k.lower()); return k

def _write_db_sheet(ws,df):
    if df.empty: return
    FLAG_COLS={"escluso_calcolo","is_listed","rating_norm","rating_below_bb",
        "rating_sp","rating_fitch","rating_moodys","rating_ifrs9"}
    C_FLAG="F4B942"
    for j,col in enumerate(df.columns,1):
        c=ws.cell(1,j,str(col))
        if col in FLAG_COLS:
            c.font=Font(name=FONT_BODY,size=SZ,bold=True,color=C_WHITE)
            c.fill=PatternFill("solid",start_color=C_FLAG)
        else:
            c.font=Font(name=FONT_BODY,size=SZ,bold=True)
        c.alignment=Alignment(vertical="center")
    for i,row in enumerate(df.itertuples(index=False),2):
        for j,val in enumerate(row,1):
            col_name=df.columns[j-1]; c=ws.cell(i,j,val)
            c.font=Font(name=FONT_BODY,size=SZ); c.alignment=Alignment(vertical="center")
            if col_name=="escluso_calcolo" and val is True:
                c.fill=PatternFill("solid",start_color="FFE0B2")
    for col in ws.columns:
        w=8
        for cell in col:
            if cell.value: w=max(w,min(len(str(cell.value))+2,40))
        ws.column_dimensions[get_column_letter(col[0].column)].width=w

def _write_legend_sheet(ws):
    _kpmg_logo(ws,"Legenda e note metodologiche")
    legenda=[("COLORE","ESITO","SIGNIFICATO"),
        ("Verde","OK","Limite rispettato"),
        ("Rosso","SFORAMENTO MAX","Valore effettivo supera il limite massimo"),
        ("Giallo","SOTTO MINIMO / VICINO","Sotto il minimo, oppure oltre l'80% del limite"),
        ("Grigio","NON RILEVABILE","Impossibile calcolare - dati mancanti o categoria non mappata"),
        ("Azzurro","ESENTE","Posizione/emittente esente dal limite (es. Titoli di Stato AAA)")]
    COLORS=[C_BLUE,C_OK,C_ERR,C_WARN,C_GRAY,C_ESENTE]
    for i,(cc,esito,desc) in enumerate(legenda,5):
        bg=COLORS[i-5]
        for j,val in enumerate([cc,esito,desc],2):
            c=ws.cell(i,j,val)
            c.font=Font(name=FONT_BODY,size=SZ,bold=(i==5),color=C_WHITE if i==5 else C_BLACK)
            c.fill=PatternFill("solid",start_color=bg); c.border=_border()
            c.alignment=Alignment(vertical="center")
        ws.row_dimensions[i].height=14
    nr=12
    ws.cell(nr,2,"Note metodologiche:").font=Font(name=FONT_BODY,size=SZ,bold=True); nr+=1
    notes=[
        "- Ogni check ha un proprio foglio di dettaglio (D_<check>): stessa base e stessa logica del check.",
        "- I fogli di dettaglio del regolamento esistono solo per le categorie effettivamente estratte dal PDF.",
        "- Concentrazione emittente: i Titoli di Stato AAA sono esenti dal limite del 10% (riga azzurra).",
        "- Giallo nei dettagli di concentrazione = emittente/gruppo oltre l'80% del proprio limite.",
        "- Denominatore: Totale attivita' / NAV dal rendiconto; fallback al totale SHIP se assente.",
        "- Rating: fallback chain S&P > Fitch > Moody's > IFRS9. Titoli azionari esenti dal rating minimo.",
    ]
    for n in notes:
        ws.cell(nr,2,n).font=Font(name=FONT_BODY,size=SZ,color=C_BLACK)
        ws.row_dimensions[nr].height=14; nr+=1
    _auto_width(ws)

# ---------------------------------------------------------------------------
def genera_excel(df_portafoglio,results_474,results_reg,limiti_regolamento,
                 info_fondo,nome_fondo="Fondo"):
    wb=Workbook()
    ws_cover=wb.active; ws_cover.title="Cover"; ws_cover.sheet_view.showGridLines=False
    ws_cover.column_dimensions["A"].width=2; ws_cover.column_dimensions["B"].width=32
    ws_cover.column_dimensions["C"].width=52
    ws_cover["B2"].value="kpmg"; ws_cover["B2"].font=Font(name=FONT_LOGO,size=18,color=C_BLUE)
    ws_cover["B2"].alignment=Alignment(horizontal="left",vertical="center"); ws_cover.row_dimensions[2].height=28
    ws_cover["B3"].value="Verifica Limiti Fondi Interni UL - Circolare ISVAP 474/D"
    ws_cover["B3"].font=Font(name=FONT_BOLD,size=14,bold=True,color=C_BLUE)
    ws_cover["B3"].alignment=Alignment(horizontal="left",vertical="center"); ws_cover.row_dimensions[3].height=22
    ws_cover["B4"].value=nome_fondo; ws_cover["B4"].font=Font(name=FONT_BODY,size=11,color=C_BLUE)
    ws_cover.row_dimensions[4].height=18; ws_cover.row_dimensions[5].height=10

    col_val="valore_mercato" if "valore_mercato" in df_portafoglio.columns else "valore_bilancio"
    excl=df_portafoglio.get("escluso_calcolo",pd.Series(False,index=df_portafoglio.index))
    tot=df_portafoglio.loc[~excl,col_val].sum() if col_val in df_portafoglio.columns else 0
    meta=[("Fondo",info_fondo.get("nome_fondo",nome_fondo)),("Tipo fondo",info_fondo.get("tipo","-")),
        ("Compagnia",info_fondo.get("compagnia","-")),
        ("Tipo prestazione",info_fondo.get("tipo_prestazione","non_previdenziale")),
        ("Data elaborazione",datetime.date.today().strftime("%d/%m/%Y")),
        ("N. posizioni",f"{len(df_portafoglio):,}"),("Totale fondo (EUR)",f"€ {tot:,.2f}"),
        ("Check 474 eseguiti",str(len(results_474))),("Check regolamento",str(len(results_reg)))]
    for r,(lbl,val) in enumerate(meta,6):
        ws_cover.cell(r,2,lbl).font=Font(name=FONT_BODY,size=9,bold=True)
        ws_cover.cell(r,3,val).font=Font(name=FONT_BODY,size=9); ws_cover.row_dimensions[r].height=16

    # Sintesi
    _write_check_sheet(wb.create_sheet("Verifica_474"),results_474,
                       "Verifica limiti - Circolare ISVAP 474/D")
    if results_reg:
        _write_check_sheet(wb.create_sheet("Verifica_Regolamento"),results_reg,
                           "Verifica limiti - Regolamento fondo")

    # UN dettaglio per check (solo headline, con dettaglio disponibile)
    used={"cover","verifica_474","verifica_regolamento","db_grezzo",
          "limiti_regolamento_raw","legenda"}
    for res in list(results_474)+list(results_reg):
        if res.check_id.endswith("_DET"): continue
        if getattr(res,"dettaglio_df",None) is None: continue
        ws=wb.create_sheet(_safe_sheet_name(res.check_id,used))
        _write_detail_sheet(ws,res)

    _write_db_sheet(wb.create_sheet("DB_Grezzo"),df_portafoglio.head(2000))

    if limiti_regolamento:
        ws_lim=wb.create_sheet("Limiti_Regolamento_Raw")
        _kpmg_logo(ws_lim,"Limiti estratti dal regolamento (Claude AI)")
        df_lim=pd.DataFrame(limiti_regolamento)
        for j,h in enumerate(df_lim.columns,2): _header_cell(ws_lim,5,j,str(h))
        for i,row in enumerate(df_lim.itertuples(index=False),6):
            for j,val in enumerate(row,2): _data_cell(ws_lim,i,j,val,stripe=(i%2==0))
        _auto_width(ws_lim)

    _write_legend_sheet(wb.create_sheet("Legenda"))

    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.read()
